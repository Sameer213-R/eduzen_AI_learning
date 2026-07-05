import re
import google.generativeai as genai
from flask import Flask, request, jsonify, render_template
import os
from dotenv import load_dotenv
from gtts import gTTS
from flask import send_from_directory
import openpyxl
import smtplib
from email.mime.text import MIMEText
import sqlite3
from datetime import datetime
from flask import session, redirect, url_for, flash
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps


# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.getenv("FLASK_SECRET_KEY", "eduzen_secret_key_123")

# Configure Gemini API with the key from environment variable
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# Define the generation configuration
generation_config = {
    "temperature": 1,
    "top_p": 0.95,
    "top_k": 40,
    "max_output_tokens": 8192,
    "response_mime_type": "text/plain",
}

# Create the Gemini model
model = genai.GenerativeModel(
    model_name="gemini-3-flash-preview",
    generation_config=generation_config,
)

# Start a chat session
chat_session = model.start_chat(history=[])

# Store the last AI response globally
last_response = {"text": ""}

# SQLite Database Setup
DB_NAME = "chatbot.db"

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    # Users table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            student_name TEXT NOT NULL,
            standard INTEGER NOT NULL,
            parent_contact TEXT NOT NULL,
            parent_email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')
    # Chat history table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS chat_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            session_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    # Quiz results table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS quiz_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            score INTEGER,
            total INTEGER,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')
    # Migrations for existing tables
    try:
        cursor.execute('ALTER TABLE chat_history ADD COLUMN session_id TEXT')
    except sqlite3.OperationalError:
        pass
    try:
        cursor.execute('ALTER TABLE chat_history ADD COLUMN user_id INTEGER')
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()

init_db()

def save_message(session_id, role, content, user_id=None):
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute('INSERT INTO chat_history (session_id, role, content, user_id) VALUES (?, ?, ?, ?)', 
                   (session_id, role, content, user_id))
    conn.commit()
    conn.close()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route("/")
def index():
    user_id = session.get('user_id')
    student_name = session.get('student_name')
    standard = session.get('standard')
    
    avg_score = 0
    if user_id:
        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute('SELECT AVG(score * 100.0 / total) FROM quiz_results WHERE user_id = ?', (user_id,))
            result = cursor.fetchone()
            if result and result[0] is not None:
                avg_score = round(result[0])
            conn.close()
        except Exception as e:
            print(f"Error fetching progress: {e}")

    return render_template("index.html", logged_in=(user_id is not None), student_name=student_name, standard=standard, avg_score=avg_score)

@app.route('/text')
@login_required
def text():
    return render_template("text.html")

@app.route('/profile')
@login_required
def profile():
    user_id = session.get('user_id')
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('SELECT student_name, standard, parent_contact, parent_email FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            user_data = {
                "name": user[0],
                "standard": user[1],
                "contact": user[2],
                "email": user[3]
            }
            return render_template('profile.html', user=user_data)
        else:
            return "User not found", 404
    except Exception as e:
        print(f"Error fetching profile: {e}")
        return "Internal Server Error", 500

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form.get('student_name')
        standard = request.form.get('standard')
        contact = request.form.get('parent_contact')
        email = request.form.get('parent_email')
        password = request.form.get('password')

        hashed_password = generate_password_hash(password)

        try:
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO users (student_name, standard, parent_contact, parent_email, password)
                VALUES (?, ?, ?, ?, ?)
            ''', (name, standard, contact, email, hashed_password))
            conn.commit()
            conn.close()
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            return "Email already exists!", 400
        except Exception as e:
            return str(e), 500
            
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('parent_email')
        password = request.form.get('password')

        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM users WHERE parent_email = ?', (email,))
        user = cursor.fetchone()
        conn.close()

        if user and check_password_hash(user[5], password):
            session['user_id'] = user[0]
            session['student_name'] = user[1]
            session['standard'] = user[2]
            session['parent_email'] = user[4]
            return redirect(url_for('index'))
        else:
            flash("Invalid email or password!", "error")
            return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/game1')
def game1():
    return render_template("game1.html")

@app.route('/game2')
def game2():
    return render_template("game2.html")

@app.route('/game3')
def game3():
    return render_template("game3.html")

@app.route('/game4')
def game4():
    return render_template("game4.html")

@app.route('/game5')
def game5():
    return render_template("game5.html")

@app.route('/pz')
def pz():
    return render_template("pz.html")

@app.route('/loader')
def loader():
    return render_template("loader.html")
# @app.route('/getdata', methods=['POST'])
# def getdata():
#     name = request.form.get('name')
#     email = request.form.get('email')
#     message = request.form.get('message')

#     if not email:  # Check if email is None or empty
#         return "Error: Email address is required!", 400

#     # Define the Excel file
#     excel_file = "form_data.xlsx"

#     try:
#         workbook = openpyxl.load_workbook(excel_file)
#         sheet = workbook.active
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.append(["name", "email", "message"])

#     sheet.append([name, email, message])
#     workbook.save(excel_file)

#     try:
#         server = smtplib.SMTP('smtp.gmail.com', 587)  # Corrected port
#         server.starttls()
#         server.login('saichechare63@gmail.com', 'tsbs dspp rjln quud')  # Use App Password instead of actual password
#         server.sendmail(
#             'saichechare63@gmail.com',
#             email,
#             f'Subject: Thank You!\n\nThank you for sending a message. You will get a reply as soon as possible...\nYour Subject: {message}'
#         )
#         server.quit()
#     except Exception as e:
#         return f"Error sending email: {e}", 500

#     return render_template('index.html')

@app.route('/gemini-response', methods=['POST'])
def get_gemini_response():
    global last_response  # Access the global variable

    try:
        user_message = request.json['message']
        session_id = request.json.get('session_id', 'default')
        user_id = session.get('user_id')
        standard = session.get('standard', 5) # Default to 5 if not found

        if not user_message:
            return jsonify({"error": "Message is required"}), 400

        # Personalized System Instruction
        personalized_message = f"You are an intelligent AI tutor for a child in standard {standard}. " \
                               f"Explain the following in an age-appropriate way for a {standard}th grade child.\n\n" \
                               f"INSTRUCTIONS:\n" \
                               f"1. Understand the user's question carefully and answer in simple/natural language.\n" \
                               f"2. IMPORTANT RULE FOR FORMATTING:\n" \
                               f"   - If your answer is SHORT (1 or 2 sentences only): Provide it as a simple paragraph.\n" \
                               f"   - If your answer is LONG (more than 2 sentences): You MUST provide it in a POINTWISE format (1., 2., 3.) where each point is on a new line.\n" \
                               f"3. For educational answers, explain at standard {standard} level and use examples.\n" \
                               f"4. End educational answers with an 'In short:' summary.\n" \
                               f"5. Maintain a friendly, helpful, and human-like conversational tone.\n" \
                               f"6. Use emojis ONLY when appropriate and very sparingly (max 1-2).\n\n" \
                               f"User Question: {user_message}"

        response = chat_session.send_message(personalized_message)
        ai_response = response.text.strip()
        formatted_response = format_response(ai_response)

        # Store in SQLite
        save_message(session_id, 'user', user_message, user_id)
        save_message(session_id, 'ai', formatted_response, user_id)

        # Store the response in a global variable
        last_response["text"] = formatted_response

        return jsonify({"response": formatted_response})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"error": "Internal Server Error"}), 500

@app.route('/get-history', methods=['GET'])
@login_required
def get_history():
    session_id = request.args.get('session_id', 'default')
    user_id = session.get('user_id')
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT role, content FROM chat_history 
            WHERE session_id = ? AND user_id = ? 
            ORDER BY timestamp ASC
        ''', (session_id, user_id))
        rows = cursor.fetchall()
        conn.close()

        history = [{"role": row[0], "content": row[1]} for row in rows]
        return jsonify(history)
    except Exception as e:
        print(f"Error fetching history: {e}")
        return jsonify([]), 500

@app.route('/get-sessions', methods=['GET'])
@login_required
def get_sessions():
    user_id = session.get('user_id')
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        # Get first message of each session as title
        cursor.execute('''
            SELECT session_id, content 
            FROM chat_history 
            WHERE user_id = ? AND id IN (SELECT MIN(id) FROM chat_history GROUP BY session_id)
            ORDER BY timestamp DESC
        ''', (user_id,))
        rows = cursor.fetchall()
        conn.close()

        sessions = [{"session_id": row[0], "title": row[1][:30] + "..." if len(row[1]) > 30 else row[1]} for row in rows]
        return jsonify(sessions)
    except Exception as e:
        print(f"Error fetching sessions: {e}")
        return jsonify([]), 500

@app.route('/clear-history', methods=['POST'])
@login_required
def clear_history():
    user_id = session.get('user_id')
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM chat_history WHERE user_id = ?', (user_id,))
        conn.commit()
        conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"Error clearing history: {e}")
        return jsonify({"error": "Internal Server Error"}), 500

def format_response(response_text):
    response_text = re.sub(r'\*\*(.*?)\*\*', r'\1', response_text)
    response_text = re.sub(r'__(.*?)__', r'\1', response_text)
    response_text = re.sub(r'\s+', ' ', response_text)

    tts = gTTS(text=response_text, lang='en', lang_check=True)
    tts.save('static/audio.mp3')

    return response_text

@app.route("/start")
def start():
    num_questions = request.args.get('num', default=5, type=int)
    
    if not last_response["text"]:
        return "No topic available for MCQ generation! Please talk to the Knowledge Buddy first.", 400

    new_message = f"Generate exactly {num_questions} multiple-choice questions with 4 options each based on our previous conversation. " \
                  f"Mark the correct answer by appending '(Correct)' to it. " \
                  f"Format each question as:\nQuestion text\nOption 1\nOption 2\nOption 3\nOption 4\n\n" \
                  f"Context: {last_response['text']}"

    try:
        # Send the request to Gemini
        response = chat_session.send_message(new_message)
        raw_mcqs = response.text.strip()

        if not raw_mcqs:
            return "Failed to generate MCQs. Try again later.", 500

        # Process MCQs to ensure correct answers are marked
        formatted_mcqs = format_mcq_output(raw_mcqs)

        # Store the new response
        last_response["text"] = formatted_mcqs

        return render_template("start.html", mcq_response=formatted_mcqs)

    except Exception as e:
        print(f"Error generating MCQs: {e}")
        return "Error generating MCQs. Please try again.", 500

def format_mcq_output(mcq_text):
    """
    Formats the MCQs properly, ensuring correct answers are marked.
    """
    # Split by double newline and filter out empty strings/whitespace
    questions = [q.strip() for q in mcq_text.split("\n\n") if q.strip()]
    formatted_questions = []

    for question in questions:
        lines = [line.strip() for line in question.split("\n") if line.strip()]
        if len(lines) < 2:
            continue  # Skip if the format is incorrect

        # Ensure the correct answer is marked
        question_text = lines[0]
        options = lines[1:]

        found_correct = False
        for i in range(len(options)):
            if "*" in options[i] or "**" in options[i]:  # Check if Gemini marks correct answers
                options[i] = options[i].replace("*", "").strip() + " (Correct)"
                found_correct = True

        if not found_correct and options:
            options[0] += " (Correct)"  # Mark the first option correct if none is marked

        formatted_questions.append("\n".join([question_text] + options))

    return "\n\n".join(formatted_questions)





@app.route('/save-quiz', methods=['POST'])
@login_required
def save_quiz():
    user_id = session.get('user_id')
    data = request.json
    score = data.get('score')
    total = data.get('total')
    
    if score is None or total is None:
        return jsonify({"error": "Invalid data"}), 400
        
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute('INSERT INTO quiz_results (user_id, score, total) VALUES (?, ?, ?)', (user_id, score, total))
        conn.commit()
        conn.close()
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"Error saving quiz: {e}")
        return jsonify({"error": "Internal Server Error"}), 500

@app.route('/getdata', methods=['POST'])
def getdata():
    parent_name = request.form.get("parent_name")
    email = request.form.get("email")
    child_age = request.form.get("child_age")
    interest = request.form.get("interest")
    message = request.form.get("message")

    if not email:
        return "Error: Email address is required!", 400

    excel_file = "form_data.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Parent Name", "Email", "Child Age", "Interest", "Message"])

    sheet.append([parent_name, email, child_age, interest, message])
    workbook.save(excel_file)

    # Send email
    try:
        msg_content = (
            f"Hello {parent_name},\n\n"
            f"Thank you for reaching out!\n"
            f"Child Age: {child_age}\n"
            f"Interest: {interest}\n"
            f"Message: {message}\n\n"
            f"We’ll get back to you soon.\n\n"
            f"Best Regards,\nTeam"
        )
        msg = MIMEText(msg_content, "plain", "utf-8")
        msg["Subject"] = "Thank You!"
        msg["From"] = "3svpvtltd@gmail.com"
        msg["To"] = email

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("3svpvtltd@gmail.com", "dfex rycf lhcf umqz")
        server.send_message(msg)  # ✅ Correct usage
        server.quit()
    except Exception as e:
        return f"Error sending email: {e}", 500

    # ✅ Render home page after email success
    return render_template("index.html")


if __name__ == '__main__':
    app.run(debug=True)